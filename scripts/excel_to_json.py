#!/usr/bin/env python3
"""
Script pour convertir le fichier Excel ministers.xlsx en JSON ministers.json
Utilise openpyxl pour lire l'Excel et json pour écrire le fichier.
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def load_excel_data(excel_path):
    """Charge les données depuis l'Excel et retourne un dict organisé."""
    wb = load_workbook(excel_path)

    # Chargement des données de base des ministres
    ministers_sheet = wb['Ministers']
    ministers = {}
    for row in ministers_sheet.iter_rows(min_row=2, values_only=True):  # Skip header
        if not row[0]: continue  # Skip empty rows
        minister_id = row[0]
        ministers[minister_id] = {
            'id': minister_id,
            'name': row[1],
            'role': row[2],
            'email': row[3] or '',
            'party': row[4],
            'photo': row[5],
            'portfolio': row[6],
            'description': row[7],
            'superiorId': row[8],
            'ministries': [],
            'biography': [],
            'collaborators': [],
            'delegates': []
        }

    # Chargement des ministères
    ministries_sheet = wb['Ministries']
    for row in ministries_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        minister_id = row[5]  # ministerId column
        if minister_id in ministers:
            ministers[minister_id]['ministries'].append({
                'id': row[0],
                'name': row[1],
                'shortName': row[2],
                'color': row[3],
                'isPrimary': row[4].lower() == 'true',
                'roleLabel': row[6]
            })

    # Chargement des biographies
    biography_sheet = wb['Biography']
    for row in biography_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        minister_id = row[0]
        if minister_id in ministers:
            ministers[minister_id]['biography'].append({
                'title': row[1],
                'organization': row[2],
                'bioSection': row[3],
                'startDate': row[4],
                'endDate': row[5],
                'eventDate': row[6],
                'eventText': row[7],
                'ongoing': row[8].lower() == 'true',
                'sortIndex': int(row[9]) if row[9] else 0
            })

    # Chargement des collaborateurs
    collaborators_sheet = wb['Collaborators']
    for row in collaborators_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]: continue
        minister_id = row[3]  # superior_id column
        if minister_id in ministers:
            ministers[minister_id]['collaborators'].append({
                'id': row[0],
                'name': row[1],
                'full_name': row[2],
                'superior_id': minister_id,
                'job_title': row[4],
                'cabinet_role': row[5],
                'cabinet_order': int(row[6]) if row[6] else 0,
                'cabinet_badge': row[7],
                'collab_grade': row[8],
                'pole_name': row[9],
                'photo_url': row[10],
                'description': row[11]
            })

    # Chargement des délégués
    delegates_sheet = wb['Delegates']
    for row in delegates_sheet.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]: continue
        minister_id = row[0]
        delegate_id = row[1]
        if minister_id in ministers:
            ministers[minister_id]['delegates'].append(delegate_id)

    return list(ministers.values())

def validate_data(ministers):
    """Validations de base des données."""
    ids = set()
    errors = []

    for minister in ministers:
        # ID unique
        if minister['id'] in ids:
            errors.append(f"ID dupliqué: {minister['id']}")
        ids.add(minister['id'])

        # Champs obligatoires
        required = ['id', 'name', 'role']
        for field in required:
            if not minister.get(field):
                errors.append(f"Champ obligatoire manquant pour {minister['id']}: {field}")

    if errors:
        print("Erreurs de validation:")
        for error in errors:
            print(f"  - {error}")
        return False

    print(f"✅ Validation réussie: {len(ministers)} ministres chargés")
    return True

def main():
    """Fonction principale."""
    excel_path = Path(__file__).parent.parent / 'data' / 'ministers.xlsx'
    json_path = Path(__file__).parent.parent / 'data' / 'ministers.json'

    if not excel_path.exists():
        print(f"❌ Fichier Excel non trouvé: {excel_path}")
        sys.exit(1)

    print(f"📖 Lecture du fichier Excel: {excel_path}")
    try:
        ministers = load_excel_data(excel_path)
    except Exception as e:
        print(f"❌ Erreur lors de la lecture de l'Excel: {e}")
        sys.exit(1)

    if not validate_data(ministers):
        sys.exit(1)

    print(f"💾 Écriture du fichier JSON: {json_path}")
    try:
        with open(json_path, 'w', encoding='utf-8') as f:
            json.dump(ministers, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"❌ Erreur lors de l'écriture du JSON: {e}")
        sys.exit(1)

    print("✅ Conversion terminée avec succès!")

if __name__ == '__main__':
    main()