#!/usr/bin/env python3
"""
Script pour remplir le fichier Excel ministers.xlsx avec les données des fichiers CSV existants.
"""

import csv
import uuid
from pathlib import Path
from openpyxl import load_workbook

def load_csv_data(csv_path):
    """Charge les données d'un fichier CSV."""
    data = []
    with open(csv_path, 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            data.append(row)
    return data

def populate_excel():
    """Remplit l'Excel avec les données des CSV."""
    # Chemins des fichiers
    excel_path = Path(__file__).parent.parent / 'data' / 'ministers.xlsx'
    persons_csv = Path(__file__).parent.parent / 'Serveur gouvernement - persons.csv'
    careers_csv = Path(__file__).parent.parent / 'Serveur gouvernement - person_careers.csv'
    ministries_csv = Path(__file__).parent.parent / 'Serveur gouvernement - ministries.csv'
    person_ministries_csv = Path(__file__).parent.parent / 'Serveur gouvernement - person_ministries.csv'

    # Charger les données CSV
    print("📖 Chargement des données CSV...")
    persons_data = load_csv_data(persons_csv)
    careers_data = load_csv_data(careers_csv)
    ministries_data = load_csv_data(ministries_csv)
    person_ministries_data = load_csv_data(person_ministries_csv)

    # Charger l'Excel
    wb = load_workbook(excel_path)

    # Nettoyer les onglets existants (garder seulement les headers)
    for sheet_name in ['Ministers', 'Ministries', 'Biography', 'Collaborators', 'Delegates']:
        if sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Garder seulement la première ligne (header)
            ws.delete_rows(2, ws.max_row)

    # Préparer les données pour les ministres
    ministers = {}
    collaborators = []
    delegates = []

    # Traiter les personnes
    for person in persons_data:
        person_id = person['id']
        role = person['role']
        full_name = person['full_name']

        if role in ['minister', 'minister-delegate']:
            # C'est un ministre (principal ou délégué)
            ministers[person_id] = {
                'id': person_id,
                'name': full_name,
                'role': 'minister',  # Normaliser à 'minister' pour le JSON
                'email': person.get('email', ''),
                'party': person.get('party', ''),
                'photo': person.get('photo_url', ''),
                'portfolio': '',  # Sera rempli depuis person_ministries
                'description': person.get('description', ''),
                'superiorId': person.get('superior_id', '')
            }
        elif role == 'collaborator':
            # C'est un collaborateur
            collaborators.append({
                'id': person_id,
                'name': full_name,
                'full_name': full_name,
                'superior_id': person.get('superior_id', ''),
                'job_title': person.get('job_title', ''),
                'cabinet_role': person.get('cabinet_role', ''),
                'cabinet_order': int(person.get('cabinet_order', 0)) if person.get('cabinet_order') else 0,
                'cabinet_badge': person.get('cabinet_badge', ''),
                'collab_grade': person.get('collab_grade', ''),
                'pole_name': person.get('pole_name', ''),
                'photo_url': person.get('photo_url', ''),
                'description': person.get('description', '')
            })

    # Traiter les liens personne-ministère pour déterminer les portfolios
    for pm in person_ministries_data:
        person_id = pm['person_id']
        ministry_id = pm['ministry_id']
        is_primary = pm['is_primary'].upper() == 'TRUE'
        role_label = pm.get('role_label', '')

        if person_id in ministers and is_primary:
            # Trouver le nom du ministère
            ministry = next((m for m in ministries_data if m['id'] == ministry_id), None)
            if ministry:
                ministers[person_id]['portfolio'] = ministry['short_name']
                # Ajouter le ministère aux données du ministre
                ministers[person_id].setdefault('ministries', []).append({
                    'id': ministry_id,
                    'name': ministry['name'],
                    'shortName': ministry['short_name'],
                    'color': ministry['color'],
                    'isPrimary': is_primary,
                    'roleLabel': role_label
                })

    # Traiter les biographies
    biographies = []
    for career in careers_data:
        person_id = career['person_id']
        if person_id in ministers:  # Seulement pour les ministres
            biographies.append({
                'ministerId': person_id,
                'title': career['title'],
                'organization': career['organisation'],
                'bioSection': career['bio_section'],
                'startDate': career.get('start_date', ''),
                'endDate': career.get('end_date', ''),
                'eventDate': career.get('event_date', ''),
                'eventText': career.get('event_text', ''),
                'ongoing': career.get('ongoing', 'FALSE').upper() == 'TRUE',
                'sortIndex': int(career.get('sort_index', 0)) if career.get('sort_index') else 0
            })

    # Remplir l'onglet Ministers
    ws_ministers = wb['Ministers']
    for minister in ministers.values():
        ws_ministers.append([
            minister['id'],
            minister['name'],
            minister['role'],
            minister['email'],
            minister['party'],
            minister['photo'],
            minister['portfolio'],
            minister['description'],
            minister['superiorId']
        ])

    # Remplir l'onglet Ministries
    ws_ministries = wb['Ministries']
    for minister in ministers.values():
        for ministry in minister.get('ministries', []):
            ws_ministries.append([
                ministry['id'],
                ministry['name'],
                ministry['shortName'],
                ministry['color'],
                str(ministry['isPrimary']).lower(),
                minister['id'],  # ministerId
                ministry['roleLabel']
            ])

    # Remplir l'onglet Biography
    ws_biography = wb['Biography']
    for bio in biographies:
        ws_biography.append([
            bio['ministerId'],
            bio['title'],
            bio['organization'],
            bio['bioSection'],
            bio['startDate'],
            bio['endDate'],
            bio['eventDate'],
            bio['eventText'],
            str(bio['ongoing']).lower(),
            bio['sortIndex']
        ])

    # Remplir l'onglet Collaborators
    ws_collaborators = wb['Collaborators']
    for collab in collaborators:
        ws_collaborators.append([
            collab['id'],
            collab['name'],
            collab['full_name'],
            collab['superior_id'],
            collab['job_title'],
            collab['cabinet_role'],
            collab['cabinet_order'],
            collab['cabinet_badge'],
            collab['collab_grade'],
            collab['pole_name'],
            collab['photo_url'],
            collab['description']
        ])

    # Onglet Delegates - pour l'instant vide, à remplir manuellement si nécessaire
    # ws_delegates = wb['Delegates']

    # Sauvegarder
    wb.save(excel_path)
    print(f"✅ Excel rempli avec {len(ministers)} ministres, {len(biographies)} biographies, {len(collaborators)} collaborateurs")
    print(f"💾 Fichier sauvegardé: {excel_path}")

if __name__ == '__main__':
    populate_excel()