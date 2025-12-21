#!/usr/bin/env python3
"""
Script pour créer un fichier Excel exemple ministers.xlsx
avec la structure définie et des données fictives.
"""

from openpyxl import Workbook
from pathlib import Path

def create_excel_example():
    """Crée un fichier Excel exemple avec la structure définie."""
    wb = Workbook()

    # Supprimer la feuille par défaut
    wb.remove(wb.active)

    # Onglet Ministers
    ws_ministers = wb.create_sheet("Ministers")
    ws_ministers.append([
        "id", "name", "role", "email", "party", "photo", "portfolio", "description", "superiorId"
    ])
    # Exemple avec Laurent Nuñez
    ws_ministers.append([
        "c23846bc-7b27-4272-9840-63b536a6a46a",
        "Laurent NUNEZ",
        "minister",
        "",
        "Renaissance",
        "https://www.info.gouv.fr/upload/media/personality/0001/15/d6c2ce019b5d09755716075a6e17e868645cd2c9.jpg",
        "Intérieur",
        "Jusqu'alors Préfet de Police de Paris, Laurent Nuñez a occupé de nombreux postes sensibles...",
        "1c71e08c-eabe-490c-82b2-262ae5df270a"  # Premier ministre
    ])

    # Onglet Ministries
    ws_ministries = wb.create_sheet("Ministries")
    ws_ministries.append([
        "id", "name", "shortName", "color", "isPrimary", "ministerId", "roleLabel"
    ])
    ws_ministries.append([
        "f46f4a1a-437a-451b-84e4-c08011e10bb0",
        "Intérieur",
        "Intérieur",
        "#1F2937",
        "true",
        "c23846bc-7b27-4272-9840-63b536a6a46a",
        "Ministre de l'Intérieur"
    ])

    # Onglet Biography
    ws_biography = wb.create_sheet("Biography")
    ws_biography.append([
        "ministerId", "title", "organization", "bioSection", "startDate", "endDate",
        "eventDate", "eventText", "ongoing", "sortIndex"
    ])
    # Quelques entrées exemple
    ws_biography.append([
        "c23846bc-7b27-4272-9840-63b536a6a46a",
        "Préfet de police de Paris",
        "Ministère de l'Intérieur",
        "Administration (ministère de l'Intérieur)",
        "2022-06-29",
        "2025-09-29",
        None,
        None,
        "false",
        3
    ])
    ws_biography.append([
        "c23846bc-7b27-4272-9840-63b536a6a46a",
        "Directeur général de la sécurité intérieure (DGSI)",
        "Ministère de l'Intérieur",
        "Administration (ministère de l'Intérieur)",
        "2017-05-30",
        "2020-05-30",
        None,
        None,
        "false",
        4
    ])

    # Onglet Collaborators
    ws_collaborators = wb.create_sheet("Collaborators")
    ws_collaborators.append([
        "id", "name", "full_name", "superior_id", "job_title", "cabinet_role",
        "cabinet_order", "cabinet_badge", "collab_grade", "pole_name", "photo_url", "description"
    ])
    ws_collaborators.append([
        "16625aab-a67c-411b-95a9-66e7f6aee7de",
        "Magali CHARBONNEAU",
        "Magali CHARBONNEAU",
        "c23846bc-7b27-4272-9840-63b536a6a46a",
        None,
        "Directrice du cabinet",
        1,
        None,
        "dircab",
        None,
        None,
        None
    ])

    # Onglet Delegates
    ws_delegates = wb.create_sheet("Delegates")
    ws_delegates.append(["ministerId", "delegateId"])
    # Exemple si Laurent Nuñez a des délégués
    # ws_delegates.append(["c23846bc-7b27-4272-9840-63b536a6a46a", "some-delegate-id"])

    # Sauvegarder
    output_path = Path(__file__).parent.parent / 'data' / 'ministers.xlsx'
    output_path.parent.mkdir(exist_ok=True)
    wb.save(output_path)
    print(f"✅ Fichier Excel exemple créé: {output_path}")

    print("\n📋 Structure créée:")
    print("- Onglet Ministers: données de base des ministres")
    print("- Onglet Ministries: ministères associés")
    print("- Onglet Biography: entrées biographiques")
    print("- Onglet Collaborators: collaborateurs de cabinet")
    print("- Onglet Delegates: relations de délégation")
    print("\n💡 Vous pouvez maintenant modifier ce fichier Excel avec vos vraies données!")

if __name__ == '__main__':
    create_excel_example()