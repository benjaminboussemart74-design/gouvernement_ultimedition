#!/usr/bin/env python3
"""
Script pour convertir un fichier Excel en JSON pour le projet Gouvernement Lecornu II.

Schéma Excel attendu :
- Onglet "Ministers" : id, name, role, email, party, photo, portfolio, description, superiorId
- Onglet "Ministries" : id, name, shortName, color, isPrimary, roleLabel, ministerId
- Onglet "Biography" : ministerId, title, organization, bioSection, startDate, endDate, eventDate, eventText, ongoing, sortIndex
- Onglet "Collaborators" : id, name, full_name, superior_id, job_title, cabinet_role, cabinet_order, cabinet_badge, collab_grade, pole_name, photo_url, description
- Onglet "Delegates" : ministerId, delegateId

Le script génère data/ministers.json avec la structure attendue par le front.
"""

import json
import sys
from pathlib import Path
from openpyxl import load_workbook

def load_excel_data(excel_path):
    """Charge les données depuis le fichier Excel."""
    wb = load_workbook(excel_path)
    
    data = {
        'ministers': [],
        'ministries': [],
        'biography': [],
        'collaborators': [],
        'delegates': []
    }
    
    # Ministers
    if 'Ministers' in wb.sheetnames:
        ws = wb['Ministers']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # id not None
                minister = dict(zip(headers, row))
                data['ministers'].append(minister)
    
    # Ministries
    if 'Ministries' in wb.sheetnames:
        ws = wb['Ministries']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                ministry = dict(zip(headers, row))
                data['ministries'].append(ministry)
    
    # Biography
    if 'Biography' in wb.sheetnames:
        ws = wb['Biography']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                bio = dict(zip(headers, row))
                data['biography'].append(bio)
    
    # Collaborators
    if 'Collaborators' in wb.sheetnames:
        ws = wb['Collaborators']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                collab = dict(zip(headers, row))
                data['collaborators'].append(collab)
    
    # Delegates
    if 'Delegates' in wb.sheetnames:
        ws = wb['Delegates']
        headers = [cell.value for cell in ws[1]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] and row[1]:
                delegate = dict(zip(headers, row))
                data['delegates'].append(delegate)
    
    return data

def build_ministers_json(data):
    """Construit la structure JSON finale à partir des données Excel."""
    ministers = {}
    
    # Initialiser les ministres
    for m in data['ministers']:
        minister_id = m['id']
        ministers[minister_id] = {
            'id': minister_id,
            'name': m['name'],
            'role': m['role'],
            'email': m.get('email', ''),
            'party': m.get('party', ''),
            'photo': m.get('photo', ''),
            'portfolio': m.get('portfolio', ''),
            'description': m.get('description', ''),
            'superiorId': m.get('superiorId'),
            'ministries': [],
            'delegates': [],
            'biography': [],
            'collaborators': []
        }
    
    # Ajouter les ministries
    for min_data in data['ministries']:
        minister_id = min_data['ministerId']
        if minister_id in ministers:
            ministers[minister_id]['ministries'].append({
                'id': min_data['id'],
                'name': min_data['name'],
                'shortName': min_data['shortName'],
                'color': min_data['color'],
                'isPrimary': min_data['isPrimary'],
                'roleLabel': min_data['roleLabel']
            })
    
    # Ajouter les biography
    for bio in data['biography']:
        minister_id = bio['ministerId']
        if minister_id in ministers:
            ministers[minister_id]['biography'].append({
                'title': bio['title'],
                'organization': bio['organization'],
                'bioSection': bio['bioSection'],
                'startDate': bio.get('startDate'),
                'endDate': bio.get('endDate'),
                'eventDate': bio.get('eventDate'),
                'eventText': bio.get('eventText'),
                'ongoing': bio.get('ongoing', False),
                'sortIndex': bio['sortIndex']
            })
    
    # Ajouter les collaborators
    for collab in data['collaborators']:
        superior_id = collab['superior_id']
        if superior_id in ministers:
            ministers[superior_id]['collaborators'].append({
                'id': collab['id'],
                'name': collab['name'],
                'full_name': collab['full_name'],
                'superior_id': superior_id,
                'job_title': collab.get('job_title'),
                'cabinet_role': collab['cabinet_role'],
                'cabinet_order': collab['cabinet_order'],
                'cabinet_badge': collab.get('cabinet_badge'),
                'collab_grade': collab['collab_grade'],
                'pole_name': collab.get('pole_name'),
                'photo_url': collab.get('photo_url'),
                'description': collab.get('description')
            })
    
    # Ajouter les delegates
    for del_data in data['delegates']:
        minister_id = del_data['ministerId']
        delegate_id = del_data['delegateId']
        if minister_id in ministers:
            ministers[minister_id]['delegates'].append(delegate_id)
    
    # Trier biography par sortIndex
    for minister in ministers.values():
        minister['biography'].sort(key=lambda x: x['sortIndex'])
    
    return list(ministers.values())

def validate_data(ministers):
    """Validations minimales."""
    ids = set()
    errors = []
    
    for minister in ministers:
        # IDs uniques
        if minister['id'] in ids:
            errors.append(f"ID dupliqué: {minister['id']}")
        ids.add(minister['id'])
        
        # Champs obligatoires
        if not minister['name']:
            errors.append(f"Ministre {minister['id']} sans nom")
        
        # Vérifier delegates existants
        for delegate_id in minister['delegates']:
            if delegate_id not in ids:
                errors.append(f"Delegate {delegate_id} inexistant pour {minister['id']}")
    
    if errors:
        print("Erreurs de validation:")
        for error in errors:
            print(f"  - {error}")
        return False
    return True

def main():
    if len(sys.argv) != 2:
        print("Usage: python excel_to_json.py <path_to_excel>")
        sys.exit(1)
    
    excel_path = Path(sys.argv[1])
    json_path = Path("data/ministers.json")
    
    if not excel_path.exists():
        print(f"Fichier Excel introuvable: {excel_path}")
        sys.exit(1)
    
    print(f"Chargement de {excel_path}...")
    data = load_excel_data(excel_path)
    
    print("Construction du JSON...")
    ministers = build_ministers_json(data)
    
    print("Validation...")
    if not validate_data(ministers):
        sys.exit(1)
    
    print(f"Sauvegarde vers {json_path}...")
    json_path.parent.mkdir(exist_ok=True)
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(ministers, f, ensure_ascii=False, indent=2)
    
    print("✅ Conversion terminée!")

if __name__ == "__main__":
    main()